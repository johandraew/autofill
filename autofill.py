#import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.edge.options import Options  # Importar la clase Options desde el módulo adecuado

# Crear opciones de Chrome
edge_options = Options()
edge_options.add_argument('--InPrivate')

# Crear instancia del navegador Chrome con las opciones
driver = webdriver.Edge(options=edge_options)
#driver = webdriver.Edge()  # Puedes cambiar 'Chrome' por el navegador que estés usando

# Abrir el archivo Excel
wb = load_workbook('archivo.xlsx')
sheet = wb.active

# Iterar sobre las filas del archivo Excel
for row in sheet.iter_rows(min_row=2, values_only=True):  # Suponiendo que los datos empiezan en la fila 2
    nombre, correo = row
    
    #Espera cinco segundos para que no se considere un bot
    # Esperar 5 segundos (ajusta según sea necesario)
    #time.sleep(5)

    # Navegar al formulario del sitio web
    driver.get('https://lu.ma/')

    # Enviar el formulario
    registrar = driver.find_element("css selector", ".btn.luma-button.flex-center.medium.primary.solid.full-width.no-icon")
    registrar.click()

    # Llenar el formulario con los datos de la fila
    elemento_nombre = driver.find_element("name", "name")
    elemento_nombre.send_keys(nombre)
    elemento_correo = driver.find_element("name", "email")
    elemento_correo.send_keys(correo)


    # Hacer clic en el botón por el tipo 'submit'
    driver.find_element("xpath", "//button[@type='submit']").click()
    time.sleep(3)
    driver.refresh()

    # Esperar algunos segundos antes de pasar a la siguiente fila (ajusta según sea necesario)
    time.sleep(5)
# Cerrar el navegador al finalizar
driver.quit()
